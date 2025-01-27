import os
import sys
from gtts import gTTS
from pydub import AudioSegment
from openpyxl import load_workbook

def read_text_from_file(file_path):
    wb = load_workbook(file_path)
    sheet = wb.active  # 활성화된 워크시트 가져오기

    i = 0 
    words = ""
    # 전체 데이터 읽기
    for row in sheet.iter_rows(values_only=True):
        if i == 0:
            i += 1
            continue
        words += f"{row[1]}|{row[2]}|"
    return words

def text_to_speech_with_custom_pause(text, output_file="output_with_pause.mp3", pause_duration=1000):
    """
    gTTS와 Pydub를 사용해 텍스트에 사용자 정의 쉬는 시간을 추가하는 함수.
    임시로 생성된 파일은 최종 파일 저장 후 삭제합니다.

    Args:
        text (str): 변환할 텍스트.
        language (str): 음성 언어 코드.
        output_file (str): 저장할 최종 음성 파일 이름.
        pause_duration (int): 단어 사이 쉬는 시간 (밀리초).

    Returns:
        None
    """
    try:
        # 텍스트를 단어로 나누기
        words = text.split("|")
        segments = []
        temp_files = []

        for i in range(len(words)):
            # 단어를 음성으로 변환
            word = words[i].strip()  # 앞뒤 공백 제거
            if word == "": continue
            # 한글 여부 확인을 위한 유니코드 범위 체크
            is_korean = any(0xAC00 <= ord(char) <= 0xD7A3 or 0x3131 <= ord(char) <= 0x318E for char in word)
            tts = gTTS(text=word, lang="ko" if is_korean else "en", slow=False)
            word_audio = f"{word}.mp3"
            tts.save(word_audio)
            temp_files.append(word_audio)

            # 음성 파일을 AudioSegment로 읽기
            segments.append(AudioSegment.from_file(word_audio))

            # 단어 사이에 무음 추가
            segments.append(AudioSegment.silent(duration=pause_duration))

        # 모든 오디오 합치기
        final_audio = sum(segments)
        final_audio.export(output_file, format="mp3")
        print(f"음성 파일이 저장되었습니다: {output_file}")

        # 임시 파일 삭제
        for temp_file in temp_files:
            if os.path.exists(temp_file):
                os.remove(temp_file)

    except Exception as e:
        print(f"오류 발생: {e}")

# 실행 예제
if __name__ == "__main__":
    # 명령행 인수로 파일 경로 받기
    if len(sys.argv) > 1:
        text_file_path = sys.argv[1]
    else:
        print("사용법: python script.py <파일경로>")
        sys.exit(1)

    # 파일에서 텍스트 읽기
    words = read_text_from_file(text_file_path)
    mp3_file_path = text_file_path.replace(".xlsx", ".mp3")
    if words: 
        text_to_speech_with_custom_pause(words, output_file = f"./file/{mp3_file_path}", pause_duration=1000)